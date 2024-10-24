import pickle
from typing import Any

import pandas as pd
import xlsxwriter
from pathlib import Path

from openpyxl.reader.excel import load_workbook


# TODO: logging
# TODO: Go back and make this pathy function signature
def save_to_computer(o, filename, folder, extension='sdm'):
    if extension.startswith('.'):
        extension = extension[1:]
    outfile = Path(folder) / f'{filename}.{extension}'

    with open(outfile, mode='wb') as f:
        # noinspection PyTypeChecker
        pickle.dump(o, f)

    print(f'SAVE SUCCESSFUL: {folder}/{filename}.{extension}')


# TODO: pathy signature
def load(name: str, folder: str) -> Any:
    extension = 'sdm'
    path = Path(folder) / f"{name}.{extension}"
    try:
        with open(path, 'rb') as pickle_in:
            return pickle.load(pickle_in)
    # TODO: is there ever a .pickle file?
    except FileNotFoundError:
        extension = 'pickle'
        path = Path(folder) / f"{name}.{extension}"
        with open(path, 'rb') as pickle_in:
            return pickle.load(pickle_in)


def load_default_model(name: str = 'Alc_vs_Non', model_type='RF'):
    base_path = Path('App/SDM/Trained_Models')
    # TODO: sdmtm (which I assume stands for SkynDataManager Trained Model) is not a standard extension for trained
    #  models - find code that creates these and use the extension for the models generated (maybe .pt?)
    extensions = ['sdmtm', 'pickle']

    model_paths = {
        'Alc_vs_Non': f"MARS2C4{model_type}_Alc_vs_Non",
        'AUD': f"MARS2C4{model_type}_AUD",
        'Binge': f"MARS2C4{model_type}_Binge",
        'worn_vs_removed': f"worn_vs_removed_{model_type}",
        'fall_duration': "fall_duration_CLN_LinearReg",
        'fall_rate': "fall_rate_CLN_LinearReg",
        'rise_duration': "rise_duration_CLN_LinearReg",
        'rise_rate': "rise_rate_CLN_LinearReg"
    }

    if name in ['worn_vs_removed', 'fall_duration', 'fall_rate', 'rise_duration', 'rise_rate']:
        model_type = 'LinReg'

    if name not in model_paths:
        raise ValueError(f'Model {name} does not exist')

    for extension in extensions:
        model_file = base_path / f"{model_paths[name]}.{extension}"
        if model_file.exists():
            with open(model_file, mode='rb') as pickle_in:
                return pickle.load(pickle_in)

    raise FileNotFoundError(f"Model '{name}' with type '{model_type}' not found.")


def get_model_summary_sheet_name(model_name: str, data_version: str) -> str:
    """Split model name at underscores, capitalize, and add a data version.
    Not sure what the purpose is but it's an easy enough refactor."""
    model_name_parts = model_name.split('_')
    model_name_new = ' '.join([part.capitalize() for part in model_name_parts])
    return f'{model_name_new} - {data_version}'


def reorder_tabs(analyses_out_folder: str, cohort_name: str) -> None:
    # This method formerly failed because _name isn't in xlsxwriter, so I wrote it in openpyxl
    # TODO: figure out if this is ever called on a non-existing excel sheet (it shouldn't be)
    file_path = Path(analyses_out_folder) / f"skyn_report_{cohort_name}.xlsx"
    workbook = load_workbook(filename=file_path)

    sheetlist = workbook.sheetnames
    sheetlist.insert(1, sheetlist.pop())

    workbook._sheets = [workbook[sheet] for sheet in sheetlist]
    workbook.save(file_path)


def merge_using_subid(sdm_results: pd.DataFrame, merge_variables: dict[str, dict[str, Any]]) -> pd.DataFrame:
    for file, info in merge_variables.items():
        df = info['df']
        subid_column = info['subid_column']
        variables = info['variables']

        data_to_add = df[[subid_column] + variables]
        sdm_results = sdm_results.merge(data_to_add, on=subid_column, how='left')
    return sdm_results
